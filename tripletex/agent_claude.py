"""Claude-powered Tripletex agent via direct Anthropic Messages API."""

import base64
import io
import json
import logging
import os
import re
import time
from datetime import date, datetime, timedelta

import requests

from prompts import SYSTEM_PROMPT
from tool_router import route_tool_call
from schema_guard import validate_and_sanitize
from spec_catalog import search_spec, validate_generic_call

log = logging.getLogger("tripletex.agent")

TASK_LOG_FILE = os.environ.get("TASK_LOG_FILE", "/opt/tripletex/task_log.jsonl")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
CLAUDE_MODEL = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6")
ANTHROPIC_API_URL = "https://api.anthropic.com/v1/messages"

MAX_TURNS = 40
WARN_TURNS_LEFT = 3


# ─── Import shared helpers from agent.py ───
from agent import (
    _pre_validate,
    _compact_response,
    _ENTITY_FIELDS,
    _extract_fields,
    _log_task,
    _ensure_bank_account,
    _exec_api,
    _validate_plan,
)


# ─── Convert Gemini tool format → Anthropic tool format ───

def _convert_gemini_type(t: str) -> str:
    mapping = {
        "STRING": "string",
        "INTEGER": "integer",
        "NUMBER": "number",
        "BOOLEAN": "boolean",
        "ARRAY": "array",
        "OBJECT": "object",
    }
    return mapping.get(t, t.lower())


def _convert_schema(gemini_schema: dict) -> dict:
    if not gemini_schema:
        return {}
    result = {}
    if "type" in gemini_schema:
        result["type"] = _convert_gemini_type(gemini_schema["type"])
    if "description" in gemini_schema:
        result["description"] = gemini_schema["description"]
    if "enum" in gemini_schema:
        result["enum"] = gemini_schema["enum"]
    if "properties" in gemini_schema:
        result["properties"] = {
            k: _convert_schema(v) for k, v in gemini_schema["properties"].items()
        }
    if "required" in gemini_schema:
        result["required"] = gemini_schema["required"]
    if "items" in gemini_schema:
        result["items"] = _convert_schema(gemini_schema["items"])
    return result


def _load_tools():
    """Load Gemini tools and convert to Anthropic tool format."""
    tools_path = os.path.join(os.path.dirname(__file__), "gen_tools_output.json")
    with open(tools_path) as f:
        data = json.load(f)
    gemini_tools = data["tools"]

    # Add internal tools
    gemini_tools.append({
        "name": "task_complete",
        "description": "Signal that all API operations for this task are done.",
        "parameters": {
            "type": "OBJECT",
            "properties": {
                "summary": {"type": "STRING", "description": "Brief summary of what was accomplished"}
            },
            "required": ["summary"],
        },
    })
    gemini_tools.append({
        "name": "search_tripletex_spec",
        "description": "Search the Tripletex API specification to find available endpoints.",
        "parameters": {
            "type": "OBJECT",
            "properties": {
                "query": {"type": "STRING", "description": "Search keywords"},
                "method_filter": {"type": "STRING", "description": "Optional: filter by HTTP method"},
            },
            "required": ["query"],
        },
    })
    gemini_tools.append({
        "name": "tripletex_api",
        "description": "Execute any Tripletex API call directly. Use search_tripletex_spec first to find the correct endpoint.",
        "parameters": {
            "type": "OBJECT",
            "properties": {
                "method": {"type": "STRING", "description": "HTTP method: GET, POST, PUT, or DELETE"},
                "path": {"type": "STRING", "description": "API endpoint path"},
                "query_params": {"type": "OBJECT", "description": "Query parameters"},
                "body": {"type": "OBJECT", "description": "Request body (JSON)"},
            },
            "required": ["method", "path"],
        },
    })
    gemini_tools.append({
        "name": "submit_plan",
        "description": "REQUIRED FIRST STEP: Submit your execution plan BEFORE making any API calls.",
        "parameters": {
            "type": "OBJECT",
            "properties": {
                "task_type": {
                    "type": "STRING",
                    "description": "Task category: customer, employee, employee_contract, department, order_invoice, order_invoice_payment, payment_existing, credit_note, travel_expense, project, project_milestone, journal_entry, correction_voucher, supplier_invoice, reminder, bank_reconciliation, fx_payment_agio, dimension, payroll, other",
                },
                "steps": {
                    "type": "ARRAY",
                    "description": "Ordered list of planned API operations",
                    "items": {
                        "type": "OBJECT",
                        "properties": {
                            "action": {"type": "STRING", "description": "Tool name or API call"},
                            "purpose": {"type": "STRING", "description": "What this step accomplishes"},
                            "key_values": {"type": "OBJECT", "description": "Important parameter values"},
                        },
                    },
                },
            },
            "required": ["task_type", "steps"],
        },
    })

    # Convert to Anthropic format
    anthropic_tools = []
    for gt in gemini_tools:
        anthropic_tools.append({
            "name": gt["name"],
            "description": gt.get("description", ""),
            "input_schema": _convert_schema(gt.get("parameters", {})),
        })
    return anthropic_tools


TOOLS = _load_tools()
log.info("Loaded %d Anthropic-format tools for Claude", len(TOOLS))


# ─── Claude API call via Anthropic Messages API ───

def _call_claude(system: str, messages: list, api_key: str, force_tool: bool = False) -> dict:
    """Call Claude via direct Anthropic Messages API."""
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    payload = {
        "model": CLAUDE_MODEL,
        "system": system,
        "messages": messages,
        "tools": TOOLS,
        "temperature": 0.1,
        "max_tokens": 8192,
    }
    if force_tool:
        payload["tool_choice"] = {"type": "tool", "name": "submit_plan"}

    for attempt in range(6):
        try:
            resp = requests.post(ANTHROPIC_API_URL, json=payload, headers=headers, timeout=180)
            if resp.status_code == 429:
                retry_after = int(resp.headers.get("retry-after", 0))
                wait = max(retry_after, 3 * (2 ** attempt))
                log.warning("Claude API 429 on attempt %d, waiting %ds...", attempt + 1, wait)
                time.sleep(wait)
                continue
            if resp.status_code >= 500:
                log.warning("Claude API %d on attempt %d, retrying...", resp.status_code, attempt + 1)
                time.sleep(2 ** attempt)
                continue
            if resp.status_code != 200:
                log.error("Claude API error %d: %s", resp.status_code, resp.text[:500])
                resp.raise_for_status()
            return resp.json()
        except requests.exceptions.ConnectionError as e:
            log.warning("Connection error on attempt %d: %s", attempt + 1, str(e)[:200])
            time.sleep(2 ** attempt)
            if attempt == 5:
                raise
        except requests.exceptions.Timeout:
            log.warning("Claude timeout on attempt %d", attempt + 1)
            if attempt == 5:
                raise
    raise RuntimeError("Claude API failed after 6 attempts")


# ─── Attachment parsing (Anthropic format) ───

def _parse_attachment(f: dict) -> list:
    """Convert attachment to Anthropic content blocks."""
    mime = f.get("mime_type", "application/octet-stream")
    filename = f.get("filename", "unknown")
    file_data = base64.b64decode(f["content_base64"])
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    # Images → Anthropic native image block
    if mime.startswith("image/"):
        return [
            {"type": "image", "source": {"type": "base64", "media_type": mime, "data": f["content_base64"]}},
            {"type": "text", "text": f"[Attached: {filename}] -- Extract ALL amounts, dates, names, and numbers exactly as written."},
        ]

    # PDF → Anthropic native document block
    if mime == "application/pdf" or ext == "pdf":
        return [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": f["content_base64"]}},
            {"type": "text", "text": f"[PDF: {filename}] -- Extract ALL amounts, dates, supplier names, invoice numbers, due dates, and line items."},
        ]

    # Excel
    if ext in ("xlsx", "xls") or "spreadsheet" in mime or "excel" in mime:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
            tables = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [str(c) if c is not None else "" for c in rows[0]]
                md = "| " + " | ".join(headers) + " |\n"
                md += "| " + " | ".join(["---"] * len(headers)) + " |\n"
                for row in rows[1:]:
                    cells = [str(c) if c is not None else "" for c in row]
                    md += "| " + " | ".join(cells) + " |\n"
                tables.append(f"### Sheet: {sheet}\n{md}")
            wb.close()
            text = "\n".join(tables)
            log.info("Parsed spreadsheet %s: %d sheets", filename, len(tables))
            return [{"type": "text", "text": f"[Spreadsheet: {filename}]\n{text[:8000]}"}]
        except Exception as e:
            log.error("Failed to parse spreadsheet %s: %s", filename, e)
            return [{"type": "text", "text": f"[Spreadsheet: {filename}] -- Could not parse. Error: {e}"}]

    # DOCX
    if ext == "docx" or "wordprocessingml" in mime:
        try:
            import docx
            doc = docx.Document(io.BytesIO(file_data))
            text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            log.info("Parsed Word doc %s: %d chars", filename, len(text))
            return [{"type": "text", "text": f"[Document: {filename}]\n{text[:8000]}"}]
        except Exception as e:
            log.error("Failed to parse docx %s: %s", filename, e)
            return [{"type": "text", "text": f"[Document: {filename}] -- Could not parse. Error: {e}"}]

    # CSV / plain text
    if ext == "csv" or mime == "text/csv":
        try:
            text = file_data.decode("utf-8", errors="replace")
            log.info("Read CSV %s: %d chars", filename, len(text))
            return [{"type": "text", "text": f"[CSV Data: {filename}]\n{text[:8000]}"}]
        except Exception:
            return [{"type": "text", "text": f"[CSV: {filename}] -- Could not decode."}]

    # Fallback
    try:
        text = file_data.decode("utf-8", errors="replace")
        return [{"type": "text", "text": f"[File: {filename}]\n{text[:8000]}"}]
    except Exception:
        return [{"type": "text", "text": f"[Attached: {filename} ({mime})] -- Binary file, cannot read contents."}]


# ─── Agent loop ───

def solve_task_sync(body: dict):
    prompt = body["prompt"]
    files = body.get("files", [])
    creds = body["tripletex_credentials"]
    base_url = creds["base_url"]
    token = creds["session_token"]
    auth = ("0", token)
    api_key = ANTHROPIC_API_KEY

    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY not set")

    _ensure_bank_account(base_url, auth)

    task_record = {
        "prompt": prompt,
        "files": [f.get("filename", "unknown") for f in files],
        "base_url": base_url,
        "api_calls": [],
        "errors": [],
        "turns": 0,
        "outcome": "unknown",
    }
    t0 = time.time()

    # Build user message
    today = date.today().isoformat()
    has_files = len(files) > 0
    file_instruction = ""
    if has_files:
        file_instruction = ("\n\nIMPORTANT: Files are attached below with data you need. "
                           "Read the data and use tools immediately. Do NOT describe what you would do -- actually DO IT. "
                           "Start by calling submit_plan, then execute each step with tool calls.")

    # Anthropic format: content blocks in user message
    user_content = [{"type": "text", "text": f"Today's date: {today}\n\nComplete this accounting task:\n\n{prompt}{file_instruction}"}]
    for f in files:
        user_content.extend(_parse_attachment(f))

    # Messages in Anthropic format (system is separate)
    messages = [
        {"role": "user", "content": user_content},
    ]

    # Agent loop
    consecutive_errors = 0
    last_error_sig = None
    auth_failed = False
    text_only_count = 0

    for turn in range(MAX_TURNS):
        log.info("Agent turn %d/%d", turn + 1, MAX_TURNS)
        force = (turn == 0)
        result = _call_claude(SYSTEM_PROMPT, messages, api_key, force_tool=force)

        # Anthropic response: content is a list of blocks
        content_blocks = result.get("content", [])
        stop_reason = result.get("stop_reason", "")

        # Append assistant message
        messages.append({"role": "assistant", "content": content_blocks})

        # Extract tool_use blocks
        tool_uses = [b for b in content_blocks if b.get("type") == "tool_use"]

        if not tool_uses:
            # Text-only response
            text_only_count += 1
            if turn < 8:
                if turn == 0:
                    nudge = ("SYSTEM: You responded with text only. This is FORBIDDEN. "
                             "You MUST call submit_plan as your FIRST action. "
                             "Parse the task, classify it, and submit your plan NOW.")
                elif turn < 3:
                    nudge = ("SYSTEM: STILL no tool calls. You are wasting turns. "
                             "Call submit_plan immediately with task_type and steps.")
                else:
                    nudge = ("You MUST use tools. Do NOT respond with only text. "
                             "Partial work earns partial credit. Use tools NOW.")
                messages.append({"role": "user", "content": [{"type": "text", "text": nudge}]})
                continue
            break

        text_only_count = 0
        done = False
        turn_had_error = False

        # Collect tool results for this turn (Anthropic requires them in one user message)
        tool_results = []

        for tu in tool_uses:
            tu_id = tu["id"]
            tool_name = tu["name"]
            args = tu.get("input", {})

            if tool_name == "task_complete":
                log.info("Task complete: %s", args.get("summary", ""))
                done = True
                task_record["outcome"] = "completed"
                task_record["summary"] = args.get("summary", "")
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tu_id,
                    "content": json.dumps({"result": "OK"}),
                })
                continue

            if tool_name == "submit_plan":
                task_type = args.get("task_type", "other")
                plan_steps = args.get("steps", [])
                approved, feedback = _validate_plan(task_type, plan_steps)
                log.info("Plan submitted: type=%s, steps=%d, approved=%s", task_type, len(plan_steps), approved)
                if not approved:
                    log.warning("Plan REJECTED: %s", feedback)
                    feedback = f"PLAN REJECTED -- YOU MUST FIX AND RESUBMIT before making any API calls.\n\n{feedback}\n\nCall submit_plan again with the corrected plan."
                task_record["plan"] = {"type": task_type, "steps": len(plan_steps), "approved": approved}
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tu_id,
                    "content": json.dumps({"result": feedback}),
                })
                continue

            if tool_name == "search_tripletex_spec":
                query = args.get("query", "")
                method_filter = args.get("method_filter")
                results = search_spec(query, method_filter)
                api_result = json.dumps(results, ensure_ascii=False)[:3000]
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tu_id,
                    "content": api_result,
                })
                continue

            if tool_name == "tripletex_api":
                method = args.get("method", "GET")
                path = args.get("path", "")
                qp = args.get("query_params") or {}
                req_body = args.get("body")

                is_valid, warnings, req_body = validate_generic_call(method, path, qp, req_body)
                for w in warnings:
                    log.warning("GENERIC_API: %s", w)

                if not is_valid:
                    api_result = f"VALIDATION ERROR: {'; '.join(warnings)}"
                else:
                    api_result = _exec_api(base_url, auth, method, path, qp if qp else None, req_body)

                call_log = {
                    "tool": tool_name, "method": method, "endpoint": path,
                    "params": qp, "body": req_body,
                    "result_snippet": api_result[:300],
                }
                if "ERROR" in api_result:
                    call_log["error"] = True
                    task_record["errors"].append(f"tripletex_api -> {method} {path}: {api_result[:200]}")
                    turn_had_error = True
                    if "HTTP 401" in api_result:
                        auth_failed = True
                    error_sig = f"tripletex_api:{path}:{api_result[:30]}"
                    if error_sig == last_error_sig:
                        consecutive_errors += 1
                    else:
                        consecutive_errors = 1
                        last_error_sig = error_sig
                else:
                    consecutive_errors = 0
                    last_error_sig = None

                task_record["api_calls"].append(call_log)
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tu_id,
                    "content": api_result,
                })
                continue

            # Standard typed tool
            try:
                args = _pre_validate(tool_name, args)
            except ValueError as ve:
                api_result = f"BLOCKED BY GUARD: {ve}"
                log.warning("Guard blocked %s: %s", tool_name, ve)
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tu_id,
                    "content": api_result,
                })
                turn_had_error = True
                continue
            method, endpoint, params, req_body = route_tool_call(tool_name, args)
            if method is None:
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tu_id,
                    "content": json.dumps({"error": f"Unknown tool: {tool_name}"}),
                })
                continue

            api_result = _exec_api(base_url, auth, method, endpoint, params, req_body)

            call_log = {
                "tool": tool_name, "method": method, "endpoint": endpoint,
                "params": params, "body": req_body,
                "result_snippet": api_result[:300],
            }
            if "ERROR" in api_result:
                call_log["error"] = True
                task_record["errors"].append(f"{tool_name} -> {method} {endpoint}: {api_result[:200]}")
                turn_had_error = True
                if "HTTP 422" in api_result:
                    api_result = (
                        f"CRITICAL API REJECTION (422): {api_result}\n\n"
                        f"STOP. DO NOT RETRY WITH THE SAME PAYLOAD. "
                        f"Read the error message above. Fix the EXACT parameter and retry."
                    )
                if "HTTP 401" in api_result:
                    auth_failed = True
                error_sig = f"{tool_name}:{endpoint}:{api_result[:30]}"
                if error_sig == last_error_sig:
                    consecutive_errors += 1
                else:
                    consecutive_errors = 1
                    last_error_sig = error_sig
            else:
                consecutive_errors = 0
                last_error_sig = None

            task_record["api_calls"].append(call_log)
            tool_results.append({
                "type": "tool_result",
                "tool_use_id": tu_id,
                "content": api_result,
            })

        # Append all tool results as one user message
        messages.append({"role": "user", "content": tool_results})

        # Auth failure bail
        if auth_failed:
            log.error("Auth failure (401) -- aborting")
            messages.append({"role": "user", "content": [{"type": "text", "text": "SYSTEM: Authentication failed (401). Call task_complete now."}]})
            result = _call_claude(SYSTEM_PROMPT, messages, api_key)
            break

        # Repeated error bail
        if consecutive_errors >= 3:
            log.warning("Same error repeated %d times -- injecting guidance", consecutive_errors)
            if consecutive_errors >= 5:
                nudge = f"SYSTEM: Same error {consecutive_errors} times. Try completely different approach or call task_complete with partial progress."
            else:
                nudge = f"SYSTEM: Same error {consecutive_errors} times. Read the error carefully -- what field is wrong? Try different value or use search_tripletex_spec."
            # Merge nudge into last user message
            messages[-1]["content"].append({"type": "text", "text": nudge})
            consecutive_errors = 0

        if done:
            break

        # Turn-limit warning
        turns_remaining = MAX_TURNS - (turn + 1)
        if turns_remaining == WARN_TURNS_LEFT and not done:
            log.warning("Turn budget warning: %d turns remaining", turns_remaining)
            messages[-1]["content"].append({"type": "text", "text": f"SYSTEM: You have {turns_remaining} turns remaining. Call task_complete NOW."})

    # Force completion if loop ended
    if task_record["outcome"] == "unknown":
        log.warning("Turn limit reached without task_complete -- forcing completion")
        task_record["outcome"] = "forced_completion_at_turn_limit"
        messages.append({"role": "user", "content": [{"type": "text", "text": "SYSTEM: Turn limit reached. Call task_complete immediately with summary."}]})
        try:
            result = _call_claude(SYSTEM_PROMPT, messages, api_key)
            for block in result.get("content", []):
                if block.get("type") == "tool_use" and block.get("name") == "task_complete":
                    task_record["outcome"] = "completed_at_limit"
                    task_record["summary"] = block.get("input", {}).get("summary", "")
                    log.info("Forced task_complete: %s", task_record["summary"])
        except Exception as e:
            log.warning("Failed forced completion call: %s", e)

    task_record["turns"] = min(turn + 1, MAX_TURNS)
    task_record["elapsed_s"] = round(time.time() - t0, 1)
    _log_task(task_record)
    log.info("Agent finished after %d turns (%.1fs)", task_record["turns"], task_record["elapsed_s"])
