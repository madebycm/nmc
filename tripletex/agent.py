"""Gemini-powered Tripletex agent with typed tools from OpenAPI spec."""

import base64
import io
import json
import logging
import os
import re
import time
from datetime import date, datetime, timedelta

import requests

from prompts import SYSTEM_PROMPT
from tool_router import route_tool_call
from schema_guard import validate_and_sanitize
from spec_catalog import search_spec, validate_generic_call

log = logging.getLogger("tripletex.agent")

TASK_LOG_FILE = os.environ.get("TASK_LOG_FILE", "/opt/tripletex/task_log.jsonl")
GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "")
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-2.5-pro")
GEMINI_URL = "https://generativelanguage.googleapis.com/v1beta/models/{}:generateContent"

MAX_TURNS = 40
WARN_TURNS_LEFT = 3  # inject "wrap up" warning when this many turns remain


# ─── Runtime autofixes ───

def _pre_validate(tool_name: str, args: dict) -> dict:
    """Deterministic fixes BEFORE routing to API. Returns corrected args."""
    args = dict(args)

    if tool_name == "post_order":
        # Auto-fill deliveryDate = orderDate if missing
        if "deliveryDate" not in args and "orderDate" in args:
            args["deliveryDate"] = args["orderDate"]
            log.info("AUTOFIX: set deliveryDate = %s", args["deliveryDate"])
        elif "deliveryDate" not in args:
            args["deliveryDate"] = date.today().isoformat()
            args.setdefault("orderDate", args["deliveryDate"])
            log.info("AUTOFIX: set deliveryDate + orderDate = %s", args["deliveryDate"])

    elif tool_name == "search_product":
        # Reject comma-separated productNumber — only first value
        pn = args.get("productNumber", "")
        if isinstance(pn, str) and "," in pn:
            first = pn.split(",")[0].strip()
            log.warning("AUTOFIX: comma in productNumber '%s' → using '%s'", pn, first)
            args["productNumber"] = first

    elif tool_name == "post_ledger_voucher":
        # Validate postings exist and balance
        postings = args.get("postings", [])
        if not postings:
            log.warning("AUTOFIX: voucher has no postings")
        elif isinstance(postings, list):
            total = sum(p.get("amountGross", 0) for p in postings if isinstance(p, dict))
            if abs(total) > 0.01:
                log.warning("AUTOFIX: voucher postings unbalanced (sum=%.2f)", total)
            voucher_date = args.get("date", date.today().isoformat())
            for i, p in enumerate(postings):
                if not isinstance(p, dict):
                    continue
                # Ensure each posting has date
                if "date" not in p:
                    p["date"] = voucher_date
                # CRITICAL: row must be >= 1 (row 0 = system-generated, causes rejection)
                if "row" not in p or p.get("row", 0) == 0:
                    p["row"] = i + 1
                    log.info("AUTOFIX: set row=%d on posting %d", i + 1, i)
                # CRITICAL: amountGrossCurrency must equal amountGross (for NOK)
                if "amountGross" in p and "amountGrossCurrency" not in p:
                    p["amountGrossCurrency"] = p["amountGross"]
                    log.info("AUTOFIX: set amountGrossCurrency=%s on posting %d",
                             p["amountGross"], i)

    elif tool_name == "post_incomingInvoice":
        # Auto-fill externalId on orderLines if missing
        order_lines = args.get("orderLines", [])
        if isinstance(order_lines, list):
            for i, line in enumerate(order_lines):
                if isinstance(line, dict) and "externalId" not in line:
                    line["externalId"] = f"line{i+1}"
                    log.info("AUTOFIX: set externalId='line%d' on orderLine %d", i+1, i)

    elif tool_name == "post_travelExpense_cost":
        # Ensure amountCurrencyIncVat is set (API requires it, model often uses 'rate' instead)
        if "amountCurrencyIncVat" not in args:
            rate = args.get("rate") or args.get("amount", 0)
            count = args.get("count", 1) or 1
            amount = rate * count if rate else 0
            if amount:
                args["amountCurrencyIncVat"] = amount
                log.info("AUTOFIX: set amountCurrencyIncVat=%s (rate=%s * count=%s)", amount, rate, count)

    elif tool_name == "post_employee":
        # Ensure valid userType — default "STANDARD" if missing/empty
        ut = args.get("userType", "")
        if not ut or ut == "0":
            args["userType"] = "STANDARD"
            log.info("AUTOFIX: set userType=STANDARD")

    elif tool_name == "search_invoice":
        # Fix: invoiceDateTo must be > invoiceDateFrom (exclusive end date)
        date_from = args.get("invoiceDateFrom", "")
        date_to = args.get("invoiceDateTo", "")
        if date_from and date_to and date_from >= date_to:
            try:
                dt = date.fromisoformat(date_to)
                args["invoiceDateTo"] = (dt + timedelta(days=1)).isoformat()
                log.info("AUTOFIX: invoiceDateTo bumped to %s (exclusive)", args["invoiceDateTo"])
            except ValueError:
                pass

    elif tool_name == "invoice_order":
        # Default sendType if not specified
        args.setdefault("sendType", "EMAIL")

    elif tool_name == "payment_invoice":
        # Sanity check: if paidAmountCurrency >> paidAmount, they're likely swapped
        # paidAmount = NOK (bank currency, larger number)
        # paidAmountCurrency = foreign currency (smaller number)
        paid = args.get("paidAmount", 0) or 0
        paid_cur = args.get("paidAmountCurrency")
        if paid_cur is not None and paid > 0 and paid_cur > 0:
            if paid_cur > paid * 1.5:
                # paidAmountCurrency is much larger than paidAmount — almost certainly swapped
                log.warning("AUTOFIX: payment_invoice paidAmount/paidAmountCurrency likely swapped "
                            "(paidAmount=%.2f, paidAmountCurrency=%.2f) — swapping", paid, paid_cur)
                args["paidAmount"] = paid_cur
                args["paidAmountCurrency"] = paid

    return args


# ─── Structured response compaction ───

# Key fields to extract per entity type
_ENTITY_FIELDS = {
    "customer": ["id", "name", "organizationNumber", "customerNumber", "supplierNumber", "email"],
    "employee": ["id", "firstName", "lastName", "email", "employeeNumber"],
    "department": ["id", "name", "departmentNumber"],
    "product": ["id", "name", "number", "priceExcludingVatCurrency", "vatType"],
    "order": ["id", "number", "customer", "orderDate", "deliveryDate"],
    "invoice": ["id", "invoiceNumber", "invoiceDate", "invoiceDueDate", "customer",
                 "currency", "amountOutstanding", "amountOutstandingTotal",
                 "amount", "amountCurrency",
                 "amountExcludingVat", "amountExcludingVatCurrency",
                 "isCredited", "isCreditNote", "isApproved"],
    "travelExpense": ["id", "title", "employee", "state", "amount"],
    "project": ["id", "name", "number", "version", "isFixedPrice", "fixedprice", "projectManager", "customer"],
    "voucher": ["id", "number", "date", "description"],
    "supplier": ["id", "name", "supplierNumber", "organizationNumber"],
    "contact": ["id", "firstName", "lastName", "email"],
    "paymentType": ["id", "description", "displayName"],
    "costCategory": ["id", "description", "displayName"],
    "account": ["id", "number", "name"],
    "employment": ["id", "employee", "startDate", "endDate", "employmentId"],
    "employmentDetails": ["id", "employment", "date", "annualSalary", "percentageOfFullTimeEquivalent",
                          "occupationCode", "employmentType", "remunerationType", "workingHoursScheme"],
    "occupationCode": ["id", "code", "nameNO"],
    "accountingDimensionName": ["id", "dimensionName", "description", "dimensionIndex", "active"],
    "accountingDimensionValue": ["id", "displayName", "dimensionIndex", "number", "active"],
}


def _extract_fields(obj: dict, fields: list) -> dict:
    """Extract specified fields from an object, including nested refs."""
    result = {}
    for f in fields:
        if f in obj:
            val = obj[f]
            # Compact nested refs to just id + name/displayName
            if isinstance(val, dict) and "id" in val:
                compact = {"id": val["id"]}
                for k in ("name", "displayName", "number", "firstName", "lastName", "code"):
                    if k in val:
                        compact[k] = val[k]
                val = compact
            result[f] = val
    return result


def _guess_entity_type(endpoint: str) -> str | None:
    """Guess entity type from endpoint path."""
    # Action endpoints that return a DIFFERENT entity type than their base path
    if "/:invoice" in endpoint:
        return "invoice"  # /order/{id}/:invoice returns Invoice
    if "/:payment" in endpoint or "/:createCreditNote" in endpoint or "/:send" in endpoint:
        return "invoice"  # /invoice/{id}/:payment etc. returns Invoice

    # Strip path params and action suffixes
    clean = re.sub(r'/\d+', '', endpoint)
    clean = re.sub(r'/:[^/]+', '', clean)
    clean = clean.strip("/")
    # Map endpoint segments to entity types
    mapping = {
        "customer": "customer", "employee": "employee", "department": "department",
        "product": "product", "order": "order", "invoice": "invoice",
        "travelExpense": "travelExpense", "travelExpense/cost": "costCategory",
        "travelExpense/paymentType": "paymentType",
        "travelExpense/costCategory": "costCategory",
        "project": "project", "ledger/voucher": "voucher",
        "ledger/account": "account", "contact": "contact",
        "invoice/paymentType": "paymentType",
        "supplier": "supplier",
        "incomingInvoice": "invoice",
        "employee/employment": "employment",
        "employee/employment/details": "employmentDetails",
        "employee/employment/occupationCode": "occupationCode",
        "ledger/accountingDimensionName": "accountingDimensionName",
        "ledger/accountingDimensionName/search": "accountingDimensionName",
        "ledger/accountingDimensionValue": "accountingDimensionValue",
        "ledger/accountingDimensionValue/search": "accountingDimensionValue",
    }
    return mapping.get(clean)


def _compact_response(endpoint: str, method: str, data) -> str:
    """Convert API response to compact structured JSON for the model."""
    if not isinstance(data, dict):
        return json.dumps(data, ensure_ascii=False)[:2000]

    entity_type = _guess_entity_type(endpoint)
    fields = _ENTITY_FIELDS.get(entity_type, [])

    # Single value response (POST/PUT)
    if "value" in data and isinstance(data["value"], dict):
        val = data["value"]
        if fields:
            compact = _extract_fields(val, fields)
        else:
            compact = val
        result = {"value": compact}
        return json.dumps(result, ensure_ascii=False)[:3000]

    # List response (GET search)
    if "values" in data and isinstance(data["values"], list):
        values = data["values"]
        count = data.get("fullResultSize", len(values))
        if fields:
            compact_values = [_extract_fields(v, fields) for v in values[:20]]
        else:
            compact_values = values[:10]
        result = {"count": count, "values": compact_values}
        return json.dumps(result, ensure_ascii=False)[:3000]

    # Fallback
    raw = json.dumps(data, ensure_ascii=False)
    return raw[:2000]


def _log_task(entry: dict):
    try:
        entry["logged_at"] = datetime.utcnow().isoformat() + "Z"
        with open(TASK_LOG_FILE, "a") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception as e:
        log.warning("Failed to write task log: %s", e)


# ─── Load typed tools from gen_tools_output.json ───

def _load_tools():
    tools_path = os.path.join(os.path.dirname(__file__), "gen_tools_output.json")
    with open(tools_path) as f:
        data = json.load(f)
    tools = data["tools"]
    # Add task_complete tool
    tools.append({
        "name": "task_complete",
        "description": "Signal that all API operations for this task are done.",
        "parameters": {
            "type": "OBJECT",
            "properties": {
                "summary": {
                    "type": "STRING",
                    "description": "Brief summary of what was accomplished",
                }
            },
            "required": ["summary"],
        },
    })
    # Spec search tool
    tools.append({
        "name": "search_tripletex_spec",
        "description": "Search the Tripletex API specification to find available endpoints. Use this BEFORE tripletex_api to discover the correct endpoint path, method, parameters, and enum values. Returns matching operations from the OpenAPI spec.",
        "parameters": {
            "type": "OBJECT",
            "properties": {
                "query": {
                    "type": "STRING",
                    "description": "Search keywords (e.g. 'employee employment salary', 'reminder invoice', 'accounting dimension')",
                },
                "method_filter": {
                    "type": "STRING",
                    "description": "Optional: filter by HTTP method (GET, POST, PUT, DELETE)",
                },
            },
            "required": ["query"],
        },
    })
    # Generic API tool
    tools.append({
        "name": "tripletex_api",
        "description": "Execute any Tripletex API call directly. Use search_tripletex_spec first to find the correct endpoint. This tool validates parameters against the OpenAPI spec before executing.",
        "parameters": {
            "type": "OBJECT",
            "properties": {
                "method": {
                    "type": "STRING",
                    "description": "HTTP method: GET, POST, PUT, or DELETE",
                },
                "path": {
                    "type": "STRING",
                    "description": "API endpoint path (e.g. '/employee/employment', '/ledger/accountingDimensionName'). Include numeric IDs directly in path for specific resources (e.g. '/employee/employment/123').",
                },
                "query_params": {
                    "type": "OBJECT",
                    "description": "Query parameters as key-value pairs",
                },
                "body": {
                    "type": "OBJECT",
                    "description": "Request body for POST/PUT calls",
                },
            },
            "required": ["method", "path"],
        },
    })
    # Plan submission tool — forces model to plan before executing
    tools.append({
        "name": "submit_plan",
        "description": "REQUIRED FIRST STEP: Submit your execution plan BEFORE making any API calls. The plan will be validated and you will receive feedback. You MUST call this tool first.",
        "parameters": {
            "type": "OBJECT",
            "properties": {
                "task_type": {
                    "type": "STRING",
                    "description": "Task category: customer, employee, employee_contract, department, order_invoice, order_invoice_payment, payment_existing, credit_note, travel_expense, project, project_milestone, journal_entry, correction_voucher, supplier_invoice, reminder, bank_reconciliation, fx_payment_agio, dimension, payroll, other",
                },
                "steps": {
                    "type": "ARRAY",
                    "description": "Ordered list of planned API operations",
                    "items": {
                        "type": "OBJECT",
                        "properties": {
                            "action": {"type": "STRING", "description": "Tool name or API call to make"},
                            "purpose": {"type": "STRING", "description": "What this step accomplishes"},
                            "key_values": {"type": "OBJECT", "description": "Important parameter values (amounts, account numbers, IDs)"},
                        },
                    },
                },
            },
            "required": ["task_type", "steps"],
        },
    })
    return tools


# ─── Plan validator ───

def _validate_plan(task_type: str, steps: list) -> tuple[bool, str]:
    """Validate a submitted plan. Returns (approved, feedback)."""
    issues = []
    step_actions = [s.get("action", "") for s in steps]

    # Check journal entry / correction voucher balance
    if task_type in ("journal_entry", "correction_voucher"):
        for s in steps:
            kv = s.get("key_values", {})
            postings = kv.get("postings", [])
            if postings and isinstance(postings, list):
                total = sum(p.get("amountGross", 0) for p in postings if isinstance(p, dict))
                if abs(total) > 0.01:
                    issues.append(f"REJECTED: Postings do not balance (sum={total:.2f}). Debit=positive, Credit=negative. Fix amounts and resubmit.")

    # Correction voucher: verify reversal logic
    if task_type == "correction_voucher":
        has_voucher = any("voucher" in a.lower() for a in step_actions)
        if not has_voucher:
            issues.append("REJECTED: Correction voucher task must include post_ledger_voucher steps to create correction entries.")
        # Check for reversal sign guidance
        for s in steps:
            kv = s.get("key_values", {})
            postings = kv.get("postings", [])
            if postings:
                for p in postings:
                    if isinstance(p, dict) and "amountGross" not in p:
                        issues.append("WARNING: Each posting must have amountGross. Positive=debit, negative=credit. To REVERSE a debit, use negative amount on the SAME account.")

    # Credit note: MUST use createCreditNote_invoice, not manual voucher
    if task_type == "credit_note":
        has_credit_tool = any("createCreditNote" in a for a in step_actions)
        if not has_credit_tool:
            issues.append("REJECTED: Credit note tasks MUST use createCreditNote_invoice tool. Do NOT create manual vouchers. Resubmit with createCreditNote_invoice(id=invoiceId).")

    # Reminder: MUST use createReminder_invoice
    if task_type == "reminder":
        has_reminder_tool = any("createReminder" in a or "createReminder_invoice" in a for a in step_actions)
        if not has_reminder_tool:
            issues.append("REJECTED: Reminder tasks MUST use createReminder_invoice tool. Do NOT create manual invoices/products for reminder fees. Resubmit with createReminder_invoice.")

    # Payment tasks: MUST use payment_invoice
    if task_type in ("order_invoice_payment", "payment_existing"):
        has_payment = any("payment_invoice" in a for a in step_actions)
        if not has_payment:
            issues.append("REJECTED: Payment tasks MUST use payment_invoice tool. Do NOT use manual vouchers for payments — vouchers do not update invoice amountOutstanding.")

    # FX payment: check field ordering
    if task_type == "fx_payment_agio":
        for s in steps:
            kv = s.get("key_values", {})
            paid = kv.get("paidAmount", 0)
            paid_cur = kv.get("paidAmountCurrency", 0)
            if paid and paid_cur and paid_cur > paid * 1.5:
                issues.append("REJECTED: paidAmount (NOK, larger) and paidAmountCurrency (foreign, smaller) appear swapped. paidAmount=NOK amount received by bank, paidAmountCurrency=foreign currency amount on invoice.")

    # Payment reversal detection: must use payment_invoice with negative amount, NOT voucher
    # Check ALL task types — reversal can be misclassified
    step_text = " ".join(s.get("purpose", "") + " " + s.get("action", "") for s in steps).lower()
    reversal_keywords = ["reverse", "storn", "tilbakefør", "zurückge", "estorn", "annul",
                         "cancel payment", "returned", "undo", "remove payment", "rückgängig",
                         "tilbakebetal", "refund", "reimburse", "rembours"]
    is_payment_reversal = any(kw in step_text for kw in reversal_keywords)
    if is_payment_reversal:
        has_payment_tool = any("payment_invoice" in a for a in step_actions)
        if not has_payment_tool:
            issues.append("REJECTED: Payment reversals MUST use payment_invoice with NEGATIVE paidAmount. Do NOT use post_ledger_voucher — vouchers do not update the invoice's amountOutstanding. Resubmit with payment_invoice(paidAmount=-AMOUNT).")
        else:
            # Verify the paidAmount is actually negative
            for s in steps:
                if "payment_invoice" in s.get("action", ""):
                    kv = s.get("key_values", {})
                    paid = kv.get("paidAmount", 0)
                    if isinstance(paid, (int, float)) and paid > 0:
                        issues.append("WARNING: Payment reversal detected but paidAmount is positive. For reversals, paidAmount MUST be NEGATIVE (e.g. -1000).")

    # Project milestone: should have put_project and order/invoice for billing
    if task_type == "project_milestone":
        has_put_project = any("put_project" in a or ("tripletex_api" in a and "project" in s.get("purpose", "").lower()) for a, s in zip(step_actions, steps))
        if not has_put_project:
            issues.append("WARNING: Project milestone tasks should use put_project to set fixedprice. Remember to include version field from search_project response.")
        has_invoice = any("invoice_order" in a or "post_order" in a for a in step_actions)
        if not has_invoice:
            issues.append("WARNING: Project milestone invoicing requires post_order + invoice_order to bill the milestone amount.")

    # Supplier invoice: check for fallback awareness
    if task_type == "supplier_invoice":
        has_incoming = any("incomingInvoice" in a for a in step_actions)
        has_voucher_fallback = any("voucher" in a.lower() for a in step_actions)
        if has_incoming and not has_voucher_fallback:
            issues.append("WARNING: post_incomingInvoice may return 403. Plan a fallback using post_ledger_voucher (debit expense account, credit account 2400 with supplier_id).")

    if issues:
        return False, "\n".join(issues)

    return True, f"APPROVED. Task type: {task_type}, {len(steps)} steps. Execute exactly as planned."


TOOLS = _load_tools()
log.info("Loaded %d typed tools", len(TOOLS))


# ─── Gemini API call ───

def _call_gemini(contents: list, api_key: str) -> dict:
    url = GEMINI_URL.format(GEMINI_MODEL) + f"?key={api_key}"
    payload = {
        "systemInstruction": {"parts": [{"text": SYSTEM_PROMPT}]},
        "contents": contents,
        "tools": [{"functionDeclarations": TOOLS}],
        "generationConfig": {
            "temperature": 0.1,
            "maxOutputTokens": 8192,
        },
    }
    for attempt in range(3):
        try:
            resp = requests.post(url, json=payload, timeout=180)
            if resp.status_code == 429 or resp.status_code >= 500:
                log.warning("Gemini %d on attempt %d, retrying...", resp.status_code, attempt + 1)
                time.sleep(2 ** attempt)
                continue
            if resp.status_code != 200:
                log.error("Gemini API error %d: %s", resp.status_code, resp.text[:500])
                resp.raise_for_status()
            return resp.json()
        except requests.exceptions.Timeout:
            log.warning("Gemini timeout on attempt %d", attempt + 1)
            if attempt == 2:
                raise
    raise RuntimeError("Gemini API failed after 3 attempts")


# ─── Tripletex API execution ───

def _exec_api(base_url: str, auth: tuple, method: str, endpoint: str,
              params: dict | None = None, body: dict | None = None) -> str:
    # Schema guard: validate body
    if body and method in ("POST", "PUT"):
        body, warnings = validate_and_sanitize(method, endpoint, params, body)
        for w in warnings:
            log.warning("GUARD: %s", w)

    url = f"{base_url}{endpoint}"
    log.info("%s %s params=%s body=%s", method, endpoint, params,
             json.dumps(body, ensure_ascii=False)[:200] if body else None)

    try:
        resp = requests.request(
            method, url, auth=auth,
            params=params,
            json=body if method in ("POST", "PUT") else None,
            timeout=60,
        )
        status = resp.status_code
        try:
            data = resp.json()
        except Exception:
            data = resp.text

        if status >= 400:
            raw = json.dumps(data, ensure_ascii=False) if isinstance(data, (dict, list)) else str(data)
            log.warning("API error %d: %s", status, raw[:500])
            hint = ""
            if status == 422:
                hint = "\n>> FIX: Read the validation error above. Correct the invalid field(s) and retry the same call."
            elif status == 403:
                hint = "\n>> FIX: This endpoint/module is not enabled. Use a fallback approach (e.g. post_ledger_voucher instead of post_incomingInvoice)."
            elif status == 404:
                hint = "\n>> FIX: Entity not found. Search again to find the correct ID before retrying."
            return f"HTTP {status} ERROR: {raw[:2000]}{hint}"

        # Compact structured response
        result = _compact_response(endpoint, method, data)
        log.info("API %d OK (%d chars compact)", status, len(result))
        return result
    except Exception as e:
        log.error("Request failed: %s", e)
        return f"REQUEST FAILED: {e}"


# ─── Attachment parsing ───

def _parse_attachment(f: dict) -> list:
    """Convert any attachment to Gemini-compatible content parts.

    Routes by MIME type + extension:
      PDF/images → native inlineData (Gemini vision)
      XLSX/XLS   → parsed to markdown table via openpyxl
      DOCX       → extracted text via python-docx
      CSV/text   → raw text
    """
    mime = f.get("mime_type", "application/octet-stream")
    filename = f.get("filename", "unknown")
    file_data = base64.b64decode(f["content_base64"])
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    # 1. VISUAL: PDF, images → native Gemini vision
    if mime.startswith("image/") or mime == "application/pdf" or ext == "pdf":
        return [
            {"inlineData": {"mimeType": mime, "data": f["content_base64"]}},
            {"text": f"[Attached: {filename}] — Extract ALL amounts, dates, names, and numbers exactly as written."},
        ]

    # 2. EXCEL: .xlsx, .xls
    if ext in ("xlsx", "xls") or "spreadsheet" in mime or "excel" in mime:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
            tables = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [str(c) if c is not None else "" for c in rows[0]]
                md = "| " + " | ".join(headers) + " |\n"
                md += "| " + " | ".join(["---"] * len(headers)) + " |\n"
                for row in rows[1:]:
                    cells = [str(c) if c is not None else "" for c in row]
                    md += "| " + " | ".join(cells) + " |\n"
                tables.append(f"### Sheet: {sheet}\n{md}")
            wb.close()
            text = "\n".join(tables)
            log.info("Parsed spreadsheet %s: %d sheets", filename, len(tables))
            return [{"text": f"[Spreadsheet: {filename}]\n{text[:8000]}"}]
        except Exception as e:
            log.error("Failed to parse spreadsheet %s: %s", filename, e)
            return [{"text": f"[Spreadsheet: {filename}] — Could not parse. Error: {e}"}]

    # 3. DOCX
    if ext == "docx" or "wordprocessingml" in mime:
        try:
            import docx
            doc = docx.Document(io.BytesIO(file_data))
            text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            log.info("Parsed Word doc %s: %d chars", filename, len(text))
            return [{"text": f"[Document: {filename}]\n{text[:8000]}"}]
        except Exception as e:
            log.error("Failed to parse docx %s: %s", filename, e)
            return [{"text": f"[Document: {filename}] — Could not parse. Error: {e}"}]

    # 4. CSV
    if ext == "csv" or mime == "text/csv":
        try:
            text = file_data.decode("utf-8", errors="replace")
            log.info("Read CSV %s: %d chars", filename, len(text))
            return [{"text": f"[CSV Data: {filename}]\n{text[:8000]}"}]
        except Exception:
            return [{"text": f"[CSV: {filename}] — Could not decode."}]

    # 5. Plain text / fallback
    try:
        text = file_data.decode("utf-8", errors="replace")
        return [{"text": f"[File: {filename}]\n{text[:8000]}"}]
    except Exception:
        return [{"text": f"[Attached: {filename} ({mime})] — Binary file, cannot read contents."}]


# ─── Bank account setup ───

def _ensure_bank_account(base_url: str, auth: tuple):
    """Ensure the company has a bank account on ledger account 1920.

    Many tasks (invoicing, payments) fail silently without this.
    """
    try:
        resp = requests.get(
            f"{base_url}/ledger/account",
            auth=auth,
            params={"number": "1920", "fields": "id,bankAccountNumber"},
            timeout=30,
        )
        if resp.status_code != 200:
            log.warning("Bank account lookup failed: %d", resp.status_code)
            return
        data = resp.json()
        values = data.get("values", [])
        if not values:
            log.warning("Ledger account 1920 not found")
            return
        acct = values[0]
        acct_id = acct.get("id")
        existing = acct.get("bankAccountNumber", "")
        if existing:
            log.info("Bank account 1920 already has number: %s", existing)
            return
        # Set a bank account number
        put_resp = requests.put(
            f"{base_url}/ledger/account/{acct_id}",
            auth=auth,
            json={"id": acct_id, "bankAccountNumber": "12345678903"},
            timeout=30,
        )
        if put_resp.status_code < 300:
            log.info("Set bank account number on 1920 (id=%s)", acct_id)
        else:
            log.warning("Failed to set bank account: %d %s", put_resp.status_code, put_resp.text[:200])
    except Exception as e:
        log.warning("Bank account setup error: %s", e)


# ─── Agent loop ───

def solve_task_sync(body: dict):
    prompt = body["prompt"]
    files = body.get("files", [])
    creds = body["tripletex_credentials"]
    base_url = creds["base_url"]
    token = creds["session_token"]
    auth = ("0", token)
    api_key = GOOGLE_API_KEY

    if not api_key:
        raise ValueError("GOOGLE_API_KEY not set")

    # Ensure bank account is set up before any task operations
    _ensure_bank_account(base_url, auth)

    task_record = {
        "prompt": prompt,
        "files": [f.get("filename", "unknown") for f in files],
        "base_url": base_url,
        "api_calls": [],
        "errors": [],
        "turns": 0,
        "outcome": "unknown",
    }
    t0 = time.time()

    # Build user message — parse all attachments through the universal pipeline
    parts = []
    for f in files:
        parts.extend(_parse_attachment(f))

    today = date.today().isoformat()
    parts.append({"text": f"Today's date: {today}\n\nComplete this accounting task:\n\n{prompt}"})
    contents = [{"role": "user", "parts": parts}]

    # Agent loop
    consecutive_errors = 0
    last_error_sig = None
    auth_failed = False

    for turn in range(MAX_TURNS):
        log.info("Agent turn %d/%d", turn + 1, MAX_TURNS)
        result = _call_gemini(contents, api_key)

        candidates = result.get("candidates", [])
        if not candidates:
            log.error("No candidates: %s", json.dumps(result)[:500])
            break

        content = candidates[0].get("content", {})
        model_parts = content.get("parts", [])
        if not model_parts:
            break

        contents.append({"role": "model", "parts": model_parts})

        fn_calls = [p for p in model_parts if "functionCall" in p]
        if not fn_calls:
            if turn < 5:
                nudge = ("You MUST use tools. Do NOT respond with only text. "
                         "If you cannot do the full task, do as much as possible — "
                         "create vouchers, search accounts, create entities. "
                         "Partial work earns partial credit. Use tools NOW.")
                contents.append({"role": "user", "parts": [{"text": nudge}]})
                continue
            break

        fn_responses = []
        done = False
        turn_had_error = False

        for fc_part in fn_calls:
            fc = fc_part["functionCall"]
            tool_name = fc["name"]
            args = fc.get("args", {})

            if tool_name == "task_complete":
                log.info("Task complete: %s", args.get("summary", ""))
                done = True
                task_record["outcome"] = "completed"
                task_record["summary"] = args.get("summary", "")
                fn_responses.append({
                    "functionResponse": {"name": tool_name, "response": {"result": "OK"}}
                })
                continue

            if tool_name == "submit_plan":
                task_type = args.get("task_type", "other")
                plan_steps = args.get("steps", [])
                approved, feedback = _validate_plan(task_type, plan_steps)
                log.info("Plan submitted: type=%s, steps=%d, approved=%s", task_type, len(plan_steps), approved)
                if not approved:
                    log.warning("Plan REJECTED: %s", feedback)
                    feedback = f"⛔ PLAN REJECTED — YOU MUST FIX AND RESUBMIT before making any API calls.\n\n{feedback}\n\nCall submit_plan again with the corrected plan. Do NOT proceed with rejected steps."
                task_record["plan"] = {"type": task_type, "steps": len(plan_steps), "approved": approved}
                fn_responses.append({
                    "functionResponse": {"name": tool_name, "response": {"result": feedback}}
                })
                continue

            if tool_name == "search_tripletex_spec":
                query = args.get("query", "")
                method_filter = args.get("method_filter")
                results = search_spec(query, method_filter)
                api_result = json.dumps(results, ensure_ascii=False)[:3000]
                fn_responses.append({
                    "functionResponse": {"name": tool_name, "response": {"result": api_result}}
                })
                continue

            if tool_name == "tripletex_api":
                method = args.get("method", "GET")
                path = args.get("path", "")
                qp = args.get("query_params") or {}
                req_body = args.get("body")

                # Validate against spec
                is_valid, warnings, req_body = validate_generic_call(method, path, qp, req_body)
                for w in warnings:
                    log.warning("GENERIC_API: %s", w)

                if not is_valid:
                    api_result = f"VALIDATION ERROR: {'; '.join(warnings)}"
                else:
                    api_result = _exec_api(base_url, auth, method, path, qp if qp else None, req_body)

                call_log = {
                    "tool": tool_name, "method": method, "endpoint": path,
                    "params": qp, "body": req_body,
                    "result_snippet": api_result[:300],
                }
                if "ERROR" in api_result:
                    call_log["error"] = True
                    task_record["errors"].append(f"tripletex_api → {method} {path}: {api_result[:200]}")
                    turn_had_error = True
                    if "HTTP 401" in api_result:
                        auth_failed = True
                    error_sig = f"tripletex_api:{path}:{api_result[:30]}"
                    if error_sig == last_error_sig:
                        consecutive_errors += 1
                    else:
                        consecutive_errors = 1
                        last_error_sig = error_sig
                else:
                    consecutive_errors = 0
                    last_error_sig = None

                task_record["api_calls"].append(call_log)
                fn_responses.append({
                    "functionResponse": {"name": tool_name, "response": {"result": api_result}}
                })
                continue

            # Runtime autofixes then route
            args = _pre_validate(tool_name, args)
            method, endpoint, params, req_body = route_tool_call(tool_name, args)
            if method is None:
                fn_responses.append({
                    "functionResponse": {"name": tool_name, "response": {"error": f"Unknown tool: {tool_name}"}}
                })
                continue

            api_result = _exec_api(base_url, auth, method, endpoint, params, req_body)

            call_log = {
                "tool": tool_name, "method": method, "endpoint": endpoint,
                "params": params, "body": req_body,
                "result_snippet": api_result[:300],
            }
            if "ERROR" in api_result:
                call_log["error"] = True
                task_record["errors"].append(f"{tool_name} → {method} {endpoint}: {api_result[:200]}")
                turn_had_error = True

                # Detect 401 auth failure — unrecoverable
                if "HTTP 401" in api_result:
                    auth_failed = True

                # Detect repeated errors (same tool+endpoint+status)
                error_sig = f"{tool_name}:{endpoint}:{api_result[:30]}"
                if error_sig == last_error_sig:
                    consecutive_errors += 1
                else:
                    consecutive_errors = 1
                    last_error_sig = error_sig
            else:
                consecutive_errors = 0
                last_error_sig = None

            task_record["api_calls"].append(call_log)

            fn_responses.append({
                "functionResponse": {"name": tool_name, "response": {"result": api_result}}
            })

        contents.append({"role": "user", "parts": fn_responses})

        # Bail on unrecoverable auth failure
        if auth_failed:
            log.error("Auth failure (401) — aborting agent loop")
            contents.append({"role": "user", "parts": [
                {"text": "SYSTEM: Authentication failed (401 Unauthorized). The API credentials are invalid. Call task_complete now and report the auth failure."}
            ]})
            # Give Gemini one more turn to call task_complete
            result = _call_gemini(contents, api_key)
            break

        # Bail on repeated identical errors (likely stuck in a loop)
        if consecutive_errors >= 3:
            log.warning("Same error repeated %d times — injecting guidance", consecutive_errors)
            if consecutive_errors >= 5:
                contents.append({"role": "user", "parts": [
                    {"text": f"SYSTEM: The same error has occurred {consecutive_errors} times. You MUST try a completely different approach now — use tripletex_api with different parameters, or use a fallback pattern. If nothing works, call task_complete with partial progress."}
                ]})
            else:
                contents.append({"role": "user", "parts": [
                    {"text": f"SYSTEM: The same error has occurred {consecutive_errors} times. Read the error carefully — what field or value is wrong? Try a different value, or use search_tripletex_spec to check the correct parameter format. Do NOT retry with identical parameters."}
                ]})
            consecutive_errors = 0  # Reset so we don't inject every turn

        if done:
            break

        # Turn-limit warning: inject "wrap up" message near the end
        turns_remaining = MAX_TURNS - (turn + 1)
        if turns_remaining == WARN_TURNS_LEFT and not done:
            log.warning("Turn budget warning: %d turns remaining", turns_remaining)
            contents.append({"role": "user", "parts": [
                {"text": f"SYSTEM: You have {turns_remaining} turns remaining. You MUST call task_complete NOW to secure partial credit for the work done so far. Summarize what you accomplished and what remains."}
            ]})

    # If loop ended without task_complete, force a completion record
    if task_record["outcome"] == "unknown":
        log.warning("Turn limit reached without task_complete — forcing completion")
        task_record["outcome"] = "forced_completion_at_turn_limit"
        # Give Gemini one final chance to summarize
        contents.append({"role": "user", "parts": [
            {"text": "SYSTEM: Turn limit reached. Call task_complete immediately with a summary of everything you accomplished."}
        ]})
        try:
            result = _call_gemini(contents, api_key)
            candidates = result.get("candidates", [])
            if candidates:
                parts = candidates[0].get("content", {}).get("parts", [])
                for p in parts:
                    fc = p.get("functionCall", {})
                    if fc.get("name") == "task_complete":
                        task_record["outcome"] = "completed_at_limit"
                        task_record["summary"] = fc.get("args", {}).get("summary", "")
                        log.info("Forced task_complete: %s", task_record["summary"])
        except Exception as e:
            log.warning("Failed forced completion call: %s", e)

    task_record["turns"] = min(turn + 1, MAX_TURNS)
    task_record["elapsed_s"] = round(time.time() - t0, 1)
    _log_task(task_record)
    log.info("Agent finished after %d turns (%.1fs)", task_record["turns"], task_record["elapsed_s"])
